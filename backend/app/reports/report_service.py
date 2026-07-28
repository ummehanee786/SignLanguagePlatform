"""
report_service.py

Task 2: once multiple Assessment Records exist (app/ai/assessment/), turn
them into a summary report:

    Total attempts, correct vs incorrect, overall score, gesture-wise
    performance, most difficult gestures, average confidence, average
    response time, improvement across attempts.

Exports as JSON (mandatory) always; PDF and Excel (bonus) are attempted
only if reportlab/openpyxl are installed - a missing optional dependency
never breaks the mandatory JSON export.

This reads from ProgressService's stored attempt log - the same data
source as the Task 2 (Weekend Sprint) dashboard - but computes different,
report-specific aggregates (most notably "improvement across attempts",
which the dashboard doesn't need). Shares _per_letter_stats() with
ProgressService rather than recomputing per-letter aggregation from
scratch.
"""

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from app.services.progress_service import ProgressService

# A letter needs at least this many attempts to be considered a
# meaningful "most difficult gesture" - same reasoning as
# MIN_ATTEMPTS_FOR_RANKING in progress_service.py (avoids one unlucky
# attempt on a rarely-practiced letter dominating the ranking).
MIN_ATTEMPTS_FOR_DIFFICULTY_RANKING = 3
MOST_DIFFICULT_TOP_N = 5


class ReportService:
    def __init__(self, progress_service: ProgressService):
        self._progress_service = progress_service

    def generate_report(self, student_id: str) -> dict:
        student_id = student_id.strip()
        # chronological order (oldest first) - needed for "improvement
        # across attempts"; get_history() returns newest-first, so reverse it.
        attempts = list(reversed(self._progress_service.get_history(student_id, limit=None)))

        total_attempts = len(attempts)
        if total_attempts == 0:
            return {
                "student_id": student_id,
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "total_assessment_attempts": 0,
                "correct_attempts": 0,
                "incorrect_attempts": 0,
                "overall_assessment_score": 0.0,
                "gesture_wise_performance": [],
                "most_difficult_gestures": [],
                "average_confidence": 0.0,
                "average_response_time_seconds": 0.0,
                "improvement_across_attempts": None,
            }

        correct_attempts = sum(1 for a in attempts if a["correct"])
        incorrect_attempts = total_attempts - correct_attempts
        overall_score = round(100 * correct_attempts / total_attempts, 2)
        average_confidence = round(sum(a["confidence"] for a in attempts) / total_attempts, 4)

        response_times = [a["time_taken_seconds"] for a in attempts if a.get("time_taken_seconds") is not None]
        average_response_time = round(sum(response_times) / len(response_times), 2) if response_times else 0.0

        per_letter = self._progress_service.per_letter_stats(attempts)
        gesture_wise_performance = sorted(
            (
                {
                    "alphabet": letter,
                    "attempts": stats["attempts"],
                    "correct": stats["correct"],
                    "accuracy_percentage": round(100 * stats["correct"] / stats["attempts"], 2),
                }
                for letter, stats in per_letter.items()
            ),
            key=lambda x: x["alphabet"],
        )

        most_difficult = sorted(
            (
                g for g in gesture_wise_performance
                if g["attempts"] >= MIN_ATTEMPTS_FOR_DIFFICULTY_RANKING
                and g["accuracy_percentage"] < 100.0
            ),
            key=lambda x: x["accuracy_percentage"],
        )[:MOST_DIFFICULT_TOP_N]

        improvement = self._compute_improvement(attempts)

        return {
            "student_id": student_id,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "total_assessment_attempts": total_attempts,
            "correct_attempts": correct_attempts,
            "incorrect_attempts": incorrect_attempts,
            "overall_assessment_score": overall_score,
            "gesture_wise_performance": gesture_wise_performance,
            "most_difficult_gestures": most_difficult,
            "average_confidence": average_confidence,
            "average_response_time_seconds": average_response_time,
            "improvement_across_attempts": improvement,
        }

    @staticmethod
    def _compute_improvement(attempts: list[dict]) -> Optional[dict]:
        """
        Splits the chronological attempt history into an earlier half
        and a later half and compares accuracy between them - a simple,
        easy-to-explain improvement signal. Needs at least 4 attempts
        (2 per half) to be meaningful; returns None below that rather
        than a misleading number from 1-2 data points.
        """
        if len(attempts) < 4:
            return None

        midpoint = len(attempts) // 2
        earlier, later = attempts[:midpoint], attempts[midpoint:]

        earlier_accuracy = round(100 * sum(1 for a in earlier if a["correct"]) / len(earlier), 2)
        later_accuracy = round(100 * sum(1 for a in later if a["correct"]) / len(later), 2)

        return {
            "earlier_half_accuracy": earlier_accuracy,
            "later_half_accuracy": later_accuracy,
            "change_percentage_points": round(later_accuracy - earlier_accuracy, 2),
            "earlier_half_attempts": len(earlier),
            "later_half_attempts": len(later),
        }

    # --- Export formats ---

    def export_json(self, student_id: str, output_path: Path) -> Path:
        """Mandatory export format."""
        report = self.generate_report(student_id)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2)
        return output_path

    def export_pdf(self, student_id: str, output_path: Path) -> Path:
        """
        Bonus export format. Raises a clear, actionable error if
        reportlab isn't installed, rather than a confusing traceback -
        this is an optional dependency, so its absence shouldn't look
        like a bug in the report logic itself.
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError as e:
            raise RuntimeError(
                "PDF export requires the 'reportlab' package. Install it with: "
                "pip install reportlab"
            ) from e

        report = self.generate_report(student_id)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        elements = [
            Paragraph(f"Assessment Report - {report['student_id']}", styles["Title"]),
            Paragraph(f"Generated: {report['generated_at']}", styles["Normal"]),
            Spacer(1, 16),
            Paragraph("Summary", styles["Heading2"]),
        ]

        summary_rows = [
            ["Total attempts", report["total_assessment_attempts"]],
            ["Correct", report["correct_attempts"]],
            ["Incorrect", report["incorrect_attempts"]],
            ["Overall score", f"{report['overall_assessment_score']}%"],
            ["Average confidence", report["average_confidence"]],
            ["Average response time (s)", report["average_response_time_seconds"]],
        ]
        if report["improvement_across_attempts"]:
            imp = report["improvement_across_attempts"]
            summary_rows.append([
                "Improvement (earlier -> later half)",
                f"{imp['earlier_half_accuracy']}% -> {imp['later_half_accuracy']}% "
                f"({imp['change_percentage_points']:+}pp)",
            ])
        elements.append(Table(summary_rows, colWidths=[220, 220]))
        elements.append(Spacer(1, 16))

        if report["gesture_wise_performance"]:
            elements.append(Paragraph("Gesture-wise Performance", styles["Heading2"]))
            gw_header = ["Letter", "Attempts", "Correct", "Accuracy %"]
            gw_rows = [gw_header] + [
                [g["alphabet"], g["attempts"], g["correct"], g["accuracy_percentage"]]
                for g in report["gesture_wise_performance"]
            ]
            gw_table = Table(gw_rows, colWidths=[80, 100, 100, 100])
            gw_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]))
            elements.append(gw_table)
            elements.append(Spacer(1, 16))

        if report["most_difficult_gestures"]:
            elements.append(Paragraph("Most Difficult Gestures", styles["Heading2"]))
            md_rows = [["Letter", "Accuracy %", "Attempts"]] + [
                [g["alphabet"], g["accuracy_percentage"], g["attempts"]]
                for g in report["most_difficult_gestures"]
            ]
            md_table = Table(md_rows, colWidths=[100, 100, 100])
            md_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]))
            elements.append(md_table)

        doc.build(elements)
        return output_path

    def export_excel(self, student_id: str, output_path: Path) -> Path:
        """Bonus export format. Same ImportError-guarding as export_pdf()."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError as e:
            raise RuntimeError(
                "Excel export requires the 'openpyxl' package. Install it with: "
                "pip install openpyxl"
            ) from e

        report = self.generate_report(student_id)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_rows = [
            ("Student", report["student_id"]),
            ("Generated", report["generated_at"]),
            ("Total attempts", report["total_assessment_attempts"]),
            ("Correct", report["correct_attempts"]),
            ("Incorrect", report["incorrect_attempts"]),
            ("Overall score (%)", report["overall_assessment_score"]),
            ("Average confidence", report["average_confidence"]),
            ("Average response time (s)", report["average_response_time_seconds"]),
        ]
        if report["improvement_across_attempts"]:
            imp = report["improvement_across_attempts"]
            summary_rows.append(("Earlier-half accuracy (%)", imp["earlier_half_accuracy"]))
            summary_rows.append(("Later-half accuracy (%)", imp["later_half_accuracy"]))
            summary_rows.append(("Change (percentage points)", imp["change_percentage_points"]))
        for row in summary_rows:
            summary_ws.append(row)
        for col in ("A", "B"):
            summary_ws.column_dimensions[col].width = 30

        gw_ws = wb.create_sheet("Gesture-wise Performance")
        gw_ws.append(["Letter", "Attempts", "Correct", "Accuracy %"])
        for cell in gw_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for g in report["gesture_wise_performance"]:
            gw_ws.append([g["alphabet"], g["attempts"], g["correct"], g["accuracy_percentage"]])

        md_ws = wb.create_sheet("Most Difficult Gestures")
        md_ws.append(["Letter", "Accuracy %", "Attempts"])
        for cell in md_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for g in report["most_difficult_gestures"]:
            md_ws.append([g["alphabet"], g["accuracy_percentage"], g["attempts"]])

        wb.save(output_path)
        return output_path


_report_service: Optional["ReportService"] = None


def get_report_service() -> "ReportService":
    global _report_service
    if _report_service is None:
        from app.services.progress_service import get_progress_service
        _report_service = ReportService(get_progress_service())
    return _report_service